#!/usr/bin/env python

"""Module that Reads and write in excell file 

Can read and write any .xls file only.
"""

__author__ = "Sumeet K. Sinha"
__credits__ = [""]
__license__ = "GPL"
__version__ = "2.0"
__maintainer__ = "Sumeet K. Sinha"
__email__ = "sumeet.kumar507@gmail.com"


import numpy as np;
from openpyxl import *;
import os.path;
import os;
from openpyxl.chart import ScatterChart, Reference, Series

class Excell:
	'It reads the microsoft excell file'

	FileName     = None;
	WorkBook     = None;
	WorkSheets   = None;

	def __str__(self):
		return "\
\n---------------------------------------\n\
Excell -> %s\n---------------------------------------\n\t\
WorkSheets:: %s\n\t\
" % (self.FileName,self.WorkSheets)

	def __init__ (self,inputFile):
		""" Define Excell class with output file

		Class Variables:
		FileName -- the hdf5.feioutput filename
		"""

		if inputFile.endswith('.xlsx'):
			print "ERROR :: File should end with .xls";
			exit();
		else:
			newFile= os.path.splitext(inputFile)[0]+".xlsx"
			if(os.path.exists(inputFile)):
				os.rename(inputFile,newFile);

		self.FileName = newFile;
		self.__initialize();


	def __initialize(self):

		self.WorkBook = Workbook(write_only=True);
		if(os.path.exists(self.FileName)):
			self.WorkBook = load_workbook(self.FileName,read_only=True,data_only=True,keep_vba=True);
		else:
			self.WorkBook.template = True
			self.WorkBook.save(self.FileName);
		self.WorkSheets = self.WorkBook.sheetnames;
		self.WorkBook = Workbook();
		return 

	def createSheet(self,SheetName):
		""" Create a Sheet in active workbook """
		self.WorkBook = Workbook(write_only=True);
		self.WorkBook = load_workbook(self.FileName,keep_vba=True);

		if SheetName in self.WorkBook.sheetnames:
		    print "createSheet:: " + SheetName + ' allready  exists'
		    return;

		self.WorkBook.create_sheet(SheetName);
		self.WorkBook.save(self.FileName);


	def col2num(self,col):
		""" Changes Column Name to Column Number  """
		import string;
		num = 0;
		for c in col:
			if c in string.ascii_letters:
				num = num * 26 + (ord(c.upper()) - ord('A')) + 1
		return num

	def reference(self,SheetName,StartRow,EndRow,StartColumn,EndColumn):
		""" returns refernce for the selected data """

		StartColumn = self.col2num(StartColumn);
		EndColumn = self.col2num(EndColumn);

		wb = Workbook();
		wb = load_workbook(self.FileName,read_only=True,data_only=True,keep_vba=True);
		ws = wb[SheetName];		
		return Reference(ws, StartColumn, StartRow, EndColumn, EndRow);


	def read(self,SheetName,StartRow,EndRow,StartColumn,EndColumn,ScaleFactor=1.0):
		""" Reads data in EXCELL for a given sheet and range """

		StartColumn = self.col2num(StartColumn);
		EndColumn   = self.col2num(EndColumn);

		WorkBookName = self.FileName;
		Data  = np.zeros([EndRow-StartRow+1,EndColumn-StartColumn+1],dtype=np.double);
		wb = Workbook();
		wb = load_workbook(WorkBookName,read_only=True,data_only=True,keep_vba=True);
		ws = wb[SheetName];

		StartRow=StartRow-1;
		EndRow=EndRow-1;
		StartColumn=StartColumn-1;
		EndColumn=EndColumn-1;

		row_number = 0;
		column_number = 0;

		###### Read WorkSheets
		for row in ws.rows:
			if (row_number<StartRow):
				row_number=row_number+1;
				continue;
			elif (row_number>EndRow):
				break;
			else:
				# print "row_number " + str(row_number)
				column_number = 0;
				for cell in row:
					if(column_number<StartColumn):
						column_number = column_number + 1;
						continue;
					elif (column_number>EndColumn):
						break;
					else:
						cellVal = cell.value;
						# print "column_number " + str(column_number);
						if(cellVal is not None ):
							data  = float(cellVal)*ScaleFactor;
							Data[row_number-StartRow][column_number-StartColumn] = data;

					column_number = column_number + 1;
			row_number=row_number+1;
		################

		return Data.flatten();

	def readColumn(self,SheetName,ColumnName,StartRow,EndRow,ScaleFactor=1.0):
		""" Reads column data in EXCELL for a given sheet """

		ColumnNumber = self.col2num(ColumnName);
		WorkBookName = self.FileName;
		Data  = np.zeros([EndRow-StartRow+1],dtype=np.double);
		wb = Workbook();
		wb = load_workbook(WorkBookName,read_only=True,data_only=True,keep_vba=True);
		ws = wb[SheetName];

		StartRow=StartRow-1;
		EndRow=EndRow-1;
		ColumnNumber=ColumnNumber-1;

		row_number = 0;
		column_number = 0;

		###### Read WorkSheets
		for row in ws.rows:
			if (row_number<StartRow):
				row_number=row_number+1;
				continue;
			elif (row_number>EndRow):
				break;
			else:
				column_number = 0;
				for cell in row:
					# print column_number
					if (column_number==ColumnNumber):
						cellVal = cell.value;
						if(cellVal is not None ):
							data  = float(cellVal)*ScaleFactor;
							Data[row_number-StartRow] = data;
						# else:
						# 	break;

					elif(column_number>ColumnNumber):
						break;
					column_number = column_number + 1;
			row_number=row_number+1;
		################

		return Data.flatten();

	# Gets the Excell data 
	def readRow(self,SheetName,RowNumber,StartColumn,EndColumn,ScaleFactor=1.0):
		""" Reads row data in EXCELL for a given sheet """

		StartColumnNumber = self.col2num(StartColumn);
		EndColumnNumber   = self.col2num(EndColumn);

		WorkBookName = self.FileName;
		Data  = np.zeros([EndColumnNumber-StartColumnNumber+1],dtype=np.double);
		wb = Workbook();
		wb = load_workbook(WorkBookName,read_only=True,data_only=True,keep_vba=True);
		ws = wb[SheetName];

		StartColumnNumber=StartColumnNumber-1;
		EndColumnNumber=EndColumnNumber-1;
		RowNumber=RowNumber-1;

		row_number = 0;
		column_number = 0;

		###### Read WorkSheets
		for row in ws.rows:
			if (row_number<RowNumber):
				row_number=row_number+1;
				continue;
			elif (row_number>RowNumber):
				break;
			else:
				column_number = 0;
				for cell in row:
					# print column_number
					if(column_number<StartColumnNumber):
						column_number = column_number + 1;
						continue;
					elif(column_number>EndColumnNumber):
						break;
					else:
						cellVal = cell.value;
						if(cellVal is not None ):
							data  = float(cell.value)*ScaleFactor;
							Data[column_number-StartColumnNumber] = data;
						# else:
						# 	break;

					column_number = column_number + 1;
			row_number=row_number+1;
		################

		return Data.flatten();

	def writeRow(self,SheetName,Row,Column,Data,DataLabel=None,ScaleFactor=1.0):
		""" Writes as row data in EXCELL in a given sheet """

		wb = Workbook(write_only=True);
		wb = load_workbook(self.FileName, keep_vba=True);
		ws = wb[SheetName];

		DataShape = Data.shape;
		DataDim   = len(DataShape);

		Column = self.col2num(Column);

		if(DataDim>2):
			print " Data must be a of dimension 2 or 1"
			return -1;
		elif(DataDim==1):
			NumRows = 1;
			NumCols = DataShape[0];
		elif(DataDim==2):
			NumRows = DataShape[0];
			NumCols = DataShape[1];

		if(DataLabel is not None):
			index = 0;
			for label in DataLabel:
				ws.cell(row=Row+index, column=Column).value = label;
				index = index +1;

		Column = Column+1;

		Data   = Data.reshape((NumRows,NumCols)); 

		for i in range(NumRows):
			for j in range(NumCols):
				ws.cell(row=Row+i, column=Column+j).value = (Data[i][j])

		wb.save(self.FileName);
		return 1;

	def writeColumn(self,SheetName,Row,Column,Data,DataLabel=None,ScaleFactor=1.0):
		""" Writes as column data in EXCELL in a given sheet """

		wb = Workbook(write_only=True);
		wb = load_workbook(self.FileName, keep_vba=True);
		ws = wb[SheetName];

		DataShape = Data.shape;
		DataDim   = len(DataShape);

		Column = self.col2num(Column);

		if(DataDim>2):
			print " Data must be a of dimension 2 or 1"
			return -1;
		elif(DataDim==1):
			NumRows = DataShape[0];
			NumCols = 1;
		elif(DataDim==2):
			NumRows = DataShape[0];
			NumCols = DataShape[1];

		if(DataLabel is not None):
			index = 0;
			for label in DataLabel:
				ws.cell(row=Row, column=Column+index).value = label;
				index = index +1;

		Row = Row+1;

		Data   = Data.reshape((NumRows,NumCols)); 

		for i in range(NumRows):
			for j in range(NumCols):
				ws.cell(row=Row+i, column=Column+j).value = (Data[i][j])

		wb.save(self.FileName);
		return 1;
	
	def __del__(self):
		newFile= os.path.splitext(self.FileName)[0]+".xls"
		os.rename(self.FileName,newFile);

	def ScatterChart(self,SheetName,title,xlabel,ylabel,position,xvalues,*yvalues):
		""" Creates a scatter line plot in excel at specified position """

		wb = Workbook(write_only=True);
		wb = load_workbook(self.FileName, keep_vba=True);

		ws = wb[SheetName];

		chart = ScatterChart()
		chart.title = title
		chart.style = 13
		chart.x_axis.title = xlabel
		chart.y_axis.title = ylabel

		for values in yvalues:
			series = Series(values, xvalues,title_from_data=False)
			chart.series.append(series)

		ws.add_chart(chart, position)
		wb.save(self.FileName);
